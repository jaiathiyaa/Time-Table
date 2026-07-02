from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from ..database import get_db
from ..models import Timetable, TimetableCell, TimeSlot, Section
from ..auth import RoleChecker

router = APIRouter(prefix="/export", tags=["Export Timetable"])

allow_read = Depends(RoleChecker(["super_admin", "admin", "hod", "faculty"]))

def parse_time_to_minutes(time_str: str) -> int:
    parts = time_str.split(":")
    hour = int(parts[0])
    minute = int(parts[1])
    if hour < 8:
        hour += 12
    return hour * 60 + minute

@router.get("/excel/section/{section_id}")
def export_section_excel(section_id: int, db: Session = Depends(get_db), current_user = allow_read):
    """
    Exports a section's timetable to a styled Excel sheet.
    """
    tt = db.query(Timetable).filter(
        Timetable.section_id == section_id,
        Timetable.is_active == True
    ).first()
    if not tt:
        raise HTTPException(status_code=404, detail="No active timetable found for this section")

    sec_name = f"{tt.section.department.code}-{tt.section.semester.semester_number}{tt.section.name}"
    ay_name = tt.academic_year.name

    # Create Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Timetable {sec_name}"

    # Set margins
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    # Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10)
    bold_cell_font = Font(name="Calibri", size=10, bold=True)
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    break_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
    theory_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    lab_fill = PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

    # Title row
    ws.merge_cells("A1:K1")
    ws["A1"] = f"TIMETABLE MANAGEMENT SYSTEM - COLLEGE OF ENGINEERING"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:K2")
    ws["A2"] = f"Section: {sec_name} | Academic Year: {ay_name} | Regulation: {tt.semester.regulation}"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Fetch time slots
    all_slots = db.query(TimeSlot).all()
    # Group time slot columns (periods and breaks)
    # We want to represent a typical day structure: Monday - Saturday
    # Days are Monday, Tuesday, etc. Let's find unique slots structure for a day
    days_in_db = sorted(list(set(ts.day for ts in all_slots)))
    
    # We can create a header based on a single day's periods (e.g. Monday)
    monday_slots = [ts for ts in all_slots if ts.day == "Monday"]
    if not monday_slots:
        # Fallback to whatever day we have
        monday_slots = [ts for ts in all_slots if ts.day == days_in_db[0]]
    
    monday_slots.sort(key=lambda ts: parse_time_to_minutes(ts.start_time))

    # Headers: Day, Period 1, Period 2, Break, ...
    headers = ["Day"]
    for ts in monday_slots:
        if ts.is_break:
            headers.append(ts.break_name)
        else:
            headers.append(f"Period {ts.period_no}\n({ts.start_time}-{ts.end_time})")

    ws.row_dimensions[4].height = 28
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Fetch cells
    cells = db.query(TimetableCell).filter(TimetableCell.timetable_id == tt.id).all()
    
    # Group cells by day and period/break
    cell_map = {} # {(day, is_break, period_no, break_name): cell}
    for c in cells:
        ts = c.time_slot
        if ts.is_break:
            cell_map[(ts.day, True, ts.period_no, ts.break_name)] = c
        else:
            cell_map[(ts.day, False, ts.period_no, None)] = c

    # Write data rows
    row_idx = 5
    for day in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]:
        # Verify if day is in database
        day_cells = [c for c in cells if c.time_slot.day == day]
        if not day_cells:
            continue
        
        ws.row_dimensions[row_idx].height = 38
        
        # Day name
        day_cell = ws.cell(row=row_idx, column=1, value=day)
        day_cell.font = bold_cell_font
        day_cell.alignment = Alignment(horizontal="center", vertical="center")
        day_cell.border = thin_border
        day_cell.fill = break_fill

        col_idx = 2
        for ts in monday_slots:
            if ts.is_break:
                # Find break slot for this day
                key = (day, True, ts.period_no, ts.break_name)
                c = cell_map.get(key)
                val = ts.break_name
                fill = break_fill
                font = bold_cell_font
            else:
                key = (day, False, ts.period_no, None)
                c = cell_map.get(key)
                if c and c.subject:
                    sub_code = c.subject.code
                    fac_code = c.faculty.code if c.faculty else ""
                    room_info = f"R:{c.classroom.room_no}" if c.classroom else (f"L:{c.laboratory.lab_name}" if c.laboratory else "")
                    val = f"{sub_code}\n{fac_code}\n{room_info}"
                    fill = lab_fill if c.subject.is_lab else theory_fill
                    font = cell_font
                else:
                    val = "Self Study"
                    fill = theory_fill
                    font = cell_font
            
            val_cell = ws.cell(row=row_idx, column=col_idx, value=val)
            val_cell.font = font
            if fill.fill_type:
                val_cell.fill = fill
            val_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            val_cell.border = thin_border
            col_idx += 1
            
        row_idx += 1

    # Adjust column widths
    for col in ws.columns:
        cell_with_attr = next((cell for cell in col if hasattr(cell, 'column_letter')), None)
        if cell_with_attr:
            ws.column_dimensions[cell_with_attr.column_letter].width = 14
    ws.column_dimensions['A'].width = 12

    # Save to memory stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"timetable_{sec_name}_{ay_name}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/pdf/section/{section_id}")
def export_section_pdf(section_id: int, db: Session = Depends(get_db), current_user = allow_read):
    """
    Exports a section's timetable to a printable PDF.
    """
    tt = db.query(Timetable).filter(
        Timetable.section_id == section_id,
        Timetable.is_active == True
    ).first()
    if not tt:
        raise HTTPException(status_code=404, detail="No active timetable found for this section")

    sec_name = f"{tt.section.department.code}-{tt.section.semester.semester_number}{tt.section.name}"
    ay_name = tt.academic_year.name

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=30,
        rightMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#1F4E79'),
        alignment=1, # Center
        spaceAfter=5
    )
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=colors.HexColor('#4A4A4A'),
        alignment=1, # Center
        spaceAfter=15
    )

    story = []
    
    # Titles
    story.append(Paragraph("AI-POWERED TIMETABLE MANAGEMENT SYSTEM", title_style))
    story.append(Paragraph(f"Section: {sec_name}   |   Academic Year: {ay_name}   |   Regulation: {tt.semester.regulation}", subtitle_style))

    # Fetch time slots
    all_slots = db.query(TimeSlot).all()
    days_in_db = sorted(list(set(ts.day for ts in all_slots)))
    
    monday_slots = [ts for ts in all_slots if ts.day == "Monday"]
    if not monday_slots:
        monday_slots = [ts for ts in all_slots if ts.day == days_in_db[0]]
    monday_slots.sort(key=lambda ts: parse_time_to_minutes(ts.start_time))

    # Headers
    headers = [Paragraph("<b>Day</b>", ParagraphStyle('HCol', fontName='Helvetica-Bold', fontSize=10, alignment=1, textColor=colors.white))]
    for ts in monday_slots:
        if ts.is_break:
            headers.append(Paragraph(f"<b>{ts.break_name}</b>", ParagraphStyle('HCol', fontName='Helvetica-Bold', fontSize=9, alignment=1, textColor=colors.white)))
        else:
            headers.append(Paragraph(f"<b>Period {ts.period_no}<br/>({ts.start_time}-{ts.end_time})</b>", ParagraphStyle('HCol', fontName='Helvetica-Bold', fontSize=8, alignment=1, textColor=colors.white)))

    data = [headers]

    # Fetch cells
    cells = db.query(TimetableCell).filter(TimetableCell.timetable_id == tt.id).all()
    cell_map = {}
    for c in cells:
        ts = c.time_slot
        if ts.is_break:
            cell_map[(ts.day, True, ts.period_no, ts.break_name)] = c
        else:
            cell_map[(ts.day, False, ts.period_no, None)] = c

    # Table styles list
    table_styles = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#BFBFBF')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#BFBFBF')),
    ]

    row_count = 1
    for day in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]:
        day_cells = [c for c in cells if c.time_slot.day == day]
        if not day_cells:
            continue
        
        row = [Paragraph(f"<b>{day}</b>", ParagraphStyle('DCol', fontName='Helvetica-Bold', fontSize=9, alignment=1))]
        
        # Color days row header
        table_styles.append(('BACKGROUND', (0, row_count), (0, row_count), colors.HexColor('#EAEAEA')))
        
        col_idx = 1
        for ts in monday_slots:
            if ts.is_break:
                row.append(Paragraph(f"<b>{ts.break_name}</b>", ParagraphStyle('BCol', fontName='Helvetica-Bold', fontSize=8, alignment=1)))
                table_styles.append(('BACKGROUND', (col_idx, row_count), (col_idx, row_count), colors.HexColor('#F2F2F2')))
            else:
                key = (day, False, ts.period_no, None)
                c = cell_map.get(key)
                if c and c.subject:
                    sub_code = c.subject.code
                    fac_code = c.faculty.code if c.faculty else ""
                    room = f"R:{c.classroom.room_no}" if c.classroom else (f"L:{c.laboratory.lab_name}" if c.laboratory else "")
                    
                    cell_text = f"<b>{sub_code}</b><br/>{fac_code}<br/>{room}"
                    row.append(Paragraph(cell_text, ParagraphStyle('Cell', fontSize=8, alignment=1)))
                    
                    # Lab color vs Theory color
                    fill_color = colors.HexColor('#EAF2F8') if c.subject.is_lab else colors.HexColor('#F2F5F8')
                    table_styles.append(('BACKGROUND', (col_idx, row_count), (col_idx, row_count), fill_color))
                else:
                    cell_text = "<b>Self Study</b>"
                    row.append(Paragraph(cell_text, ParagraphStyle('Cell', fontSize=8, alignment=1)))
                    table_styles.append(('BACKGROUND', (col_idx, row_count), (col_idx, row_count), colors.HexColor('#F2F5F8')))
            col_idx += 1
        
        data.append(row)
        row_count += 1

    # Widths: Day: 60, rest: equal distribution
    col_widths = [60] + [65] * len(monday_slots)
    
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle(table_styles))
    
    story.append(t)
    doc.build(story)
    
    buffer.seek(0)
    filename = f"timetable_{sec_name}_{ay_name}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
